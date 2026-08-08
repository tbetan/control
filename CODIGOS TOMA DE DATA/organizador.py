import os
import glob
import re
import pandas as pd
from openpyxl import Workbook

# Carpeta donde está este script
carpeta = os.path.dirname(os.path.abspath(__file__))

# Buscar todos los Excel
archivos = glob.glob(os.path.join(carpeta, "*.xlsx"))

# Excluir el archivo de salida si existe
archivos = [a for a in archivos if "Todas_las_Datas" not in os.path.basename(a)]

# Ordenar por número de prueba (01,02,...20)
def numero_prueba(nombre):
    m = re.search(r'Prueba_(\d+)', os.path.basename(nombre))
    return int(m.group(1)) if m else 999

archivos.sort(key=numero_prueba)

wb = Workbook()
ws = wb.active
ws.title = "Todas las Datas"

for i, archivo in enumerate(archivos):

    print(f"Leyendo: {os.path.basename(archivo)}")

    df = pd.read_excel(archivo)

    # Cada DATA ocupa 4 columnas (3 de datos + 1 vacía)
    col_inicio = i * 4 + 1

    # Título
    ws.cell(row=1, column=col_inicio).value = f"DATA {i+1}"

    # Encabezados
    for c, nombre in enumerate(df.columns):
        ws.cell(row=2, column=col_inicio + c).value = nombre

    # Datos
    for f, (_, row) in enumerate(df.iterrows(), start=3):
        for c, valor in enumerate(row):
            ws.cell(row=f, column=col_inicio + c).value = valor

salida = os.path.join(carpeta, "Todas_las_Datas.xlsx")
wb.save(salida)

print("\nArchivo creado correctamente.")
print(salida)